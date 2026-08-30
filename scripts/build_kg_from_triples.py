# -*- coding: utf-8 -*-
"""
从 doc/知识图谱/【20260623】三元组新修版本.xlsx 重建 data/kg_graph_vis.json。
关系中文标签来自【20260623】知识图谱架构新修.xlsx「总表」；
节点类型来自架构表「最终级实体内容」的精确分类（优于名称启发式）。
输出格式与现有 kg_graph_vis.json 一致。
"""
import json
import os
import re
from collections import defaultdict, OrderedDict

import openpyxl

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TRIPLES = os.path.join(BASE, 'doc', '知识图谱', '【20260623】三元组新修版本.xlsx')
ARCH = os.path.join(BASE, 'doc', '知识图谱', '【20260623】知识图谱架构新修.xlsx')
OUT = os.path.join(BASE, 'data', 'kg_graph_vis.json')

TYPE_PRIORITY = ['疾病', '药物', '症状', '并发症', '不良反应', '干预方法',
                 '检查', '人群', '证候诊断', '疾病阶段', '传播途径', '其他']

# 关系 -> (subject端类型, object端类型)，用于实体类型表未覆盖时的语义推断兜底
RELATION_NODE_TYPES = {
    'contains': ('疾病', '症状'),
    'has_symptom': ('疾病', '症状'),
    'associated_with': ('症状', '症状'),
    'recommends_drug': ('疾病', '药物'),
    'discourages_drug': ('疾病', '药物'),
    'cautions_drug': ('疾病', '药物'),
    'prohibits_drug': ('疾病', '药物'),
    'causes_adverse_reaction': ('药物', '不良反应'),
    'recommends_treatment': ('疾病', '干预方法'),
    'discourages_treatment': ('疾病', '干预方法'),
    'has_specific_regimen': ('疾病', '干预方法'),
    'has_examination_method': ('症状', '检查'),
    'has_detection_target': ('检查', '检查'),
    'needs_monitoring': ('药物', '检查'),
    'has_complication': ('疾病', '并发症'),
    'has_susceptible_population': ('疾病', '人群'),
    'has_transmission_route': ('疾病', '传播途径'),
    'occurs_at_stage': ('疾病', '疾病阶段'),
    'has_syndrome_diagnosis': ('疾病', '证候诊断'),
    'recommends_use_time': ('药物', '其他'),
}


def clean(s):
    return str(s).strip() if s is not None else ''


def load_meta():
    wb = openpyxl.load_workbook(ARCH, read_only=True, data_only=True)
    ws = wb['总表']
    relations = {}
    for r in ws.iter_rows(min_row=3, values_only=True):
        rid, rlabel = r[5], r[6]
        rid = clean(rid)
        if re.match(r'^[a-z_]+$', rid):
            relations[rid] = clean(rlabel) or rid

    ws2 = wb['最终级实体内容']
    col_type = {1: '疾病', 2: '症状', 3: '药物', 4: '检查', 5: '干预方法',
                6: '人群', 7: '疾病阶段', 8: '并发症', 9: '传播途径', 10: '不良反应', 11: '证候诊断'}
    ent_type = {}
    for r in ws2.iter_rows(min_row=2, values_only=True):
        for ci, tp in col_type.items():
            v = clean(r[ci]) if ci < len(r) else ''
            if v:
                # 同名实体若出现在多列，按类型优先级保留更高的
                if v not in ent_type or TYPE_PRIORITY.index(tp) < TYPE_PRIORITY.index(ent_type[v]):
                    ent_type[v] = tp
    return relations, ent_type


def main():
    relations, ent_type = load_meta()

    wb = openpyxl.load_workbook(TRIPLES, read_only=True, data_only=True)
    ws = wb['Sheet1']

    triples = []
    relation_counts = defaultdict(int)
    header = True
    for row in ws.iter_rows(values_only=True):
        if header:
            header = False
            continue
        subj, rel, obj = clean(row[0]), clean(row[1]), clean(row[2])
        if not subj or not rel or not obj:
            continue
        if rel not in relations:
            continue  # 丢弃定义表外的脏关系
        # 丢弃实体名里混入了英文关系名的脏数据（如「精神causes_adverse_reaction」）
        if any(r in subj or r in obj for r in relations):
            continue
        triples.append((subj, rel, obj))
        relation_counts[rel] += 1

    # 去重边
    seen, edges = set(), []
    in_deg, out_deg = defaultdict(int), defaultdict(int)
    for subj, rel, obj in triples:
        key = (subj, rel, obj)
        if key in seen:
            continue
        seen.add(key)
        edges.append(OrderedDict([
            ('id', 'e%d' % len(edges)),
            ('source', subj), ('target', obj),
            ('relation', rel), ('relation_label', relations[rel]),
        ]))
        out_deg[subj] += 1
        in_deg[obj] += 1

    all_nodes = set()
    for subj, rel, obj in triples:
        all_nodes.add(subj)
        all_nodes.add(obj)

    # 关系语义推断的候选类型（实体类型表未覆盖时兜底）
    inferred = defaultdict(set)
    for subj, rel, obj in triples:
        st, ot = RELATION_NODE_TYPES.get(rel, ('其他', '其他'))
        inferred[subj].add(st)
        inferred[obj].add(ot)

    def pick_inferred(cands):
        for t in TYPE_PRIORITY:
            if t in cands:
                return t
        return '其他'

    # 名称启发式：实体类型表未收录时，按名称关键词判类型（优先于关系推断，
    # 修正"耐药检测/中成药/受体激动药"等被 contains 关系误推成疾病的节点）
    def name_type(name):
        # 中医证候：以「证」结尾，或含辨证论治
        if name.endswith('证') or '辨证' in name or '辩证' in name:
            return '证候诊断'
        kw = [
            ('检查', ['检测', '检查', '筛查', '量表', '问卷', '评估', '测量', '计数', '载量', '监测']),
            ('药物', ['汤', '丸', '散', '胶囊', '颗粒', '口服液', '合剂', '冲剂', '药', '剂', '素', '中成药', '激动药', '抑制剂', '阻滞剂', '漱口液', '注射液', '片']),
            ('干预方法', ['疗法', '治疗', '干预', '手术', '通气', '训练', '管理', '教育', '按摩', '针灸', '瑜伽', '太极', '论治']),
            ('人群', ['患者', '人群', '孕妇', '儿童', '感染者', '囚犯', '青少年', '同性恋', '使用者']),
            ('传播途径', ['传播']),
            ('疾病阶段', ['期', '阶段', '产后', '怀孕']),
        ]
        for tp, words in kw:
            for w in words:
                if w in name:
                    return tp
        return None

    nodes, type_counts = [], defaultdict(int)
    for nid in sorted(all_nodes):
        ptype = ent_type.get(nid)            # 优先：实体类型表精确分类
        if not ptype:
            ptype = name_type(nid)           # 其次：名称启发式
        if not ptype:
            ptype = pick_inferred(inferred.get(nid, set()))  # 兜底：关系语义推断
        type_counts[ptype] += 1
        nodes.append(OrderedDict([
            ('id', nid), ('label', nid), ('primary_type', ptype),
            ('in_degree', in_deg.get(nid, 0)), ('out_degree', out_deg.get(nid, 0)),
        ]))

    relation_definitions = [
        OrderedDict([('id', rel), ('label', relations[rel]), ('count', cnt)])
        for rel, cnt in sorted(relation_counts.items(), key=lambda x: -x[1])
    ]

    out = OrderedDict([
        ('nodes', nodes), ('edges', edges),
        ('relation_definitions', relation_definitions),
        ('stats', OrderedDict([
            ('node_count', len(nodes)), ('edge_count', len(edges)),
            ('type_counts', dict(type_counts)),
        ])),
    ])
    with open(OUT, 'w', encoding='utf-8') as f:
        json.dump(out, f, ensure_ascii=False, separators=(',', ':'))

    print('节点数:', len(nodes), '边数:', len(edges), '关系类型:', len(relation_definitions))
    print('类型分布:', dict(type_counts))
    print('归为「其他」的节点数:', type_counts.get('其他', 0))
    print('输出:', OUT)


if __name__ == '__main__':
    main()
